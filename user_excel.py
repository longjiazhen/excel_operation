import datetime
import MySQLdb
from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Color

class ExcelUtils(object):
    """ pip install openpyxl
        pip install pillow """
    def __init__(self):
        # 新建工作簿
        self.wb = Workbook()
        # 默认的第一个表单Sheet
        self.ws = self.wb.active
        # 新建第二个表单，命名为我的表单
        self.ws_two = self.wb.create_sheet('我的表单')
        # 将第一个表单的标题改成小可爱的表单
        self.ws.title = '小可爱的表单'
        # 将第一个表单的颜色改成红色
        self.ws.sheet_properties.tabColor = 'ff0000'
        # 新建第三个表单，不命名，默认为Sheet1
        self.ws_three = self.wb.create_sheet()

    def do_sth(self):
        # 插入数据
        self.ws['A1'] = 66
        self.ws['A2'] = '你好'

        for row in self.ws_two['A1:E5']:
            for cell in row:
                cell.value = 2

        #对数据进行求和
        self.ws_two['G1'] = '=SUM(A1:E1)'

        # 插入当前时间
        self.ws['A3'] = datetime.datetime.now()
        # 设置字体大小和颜色
        font = Font(sz=18, color='000FFF')
        self.ws['A2'].font = font
        # 插入图片
        img = Image('./static/temp.jpg')
        # 改变图片大小
        img.newSize = (360, 280)
        img.width, img.height = img.newSize
        self.ws.add_image(img, 'C1')
        # 合并单元格
        self.ws.merge_cells('H1:K2')
        # 取消合并
        # self.ws.unmerge_cells('H1:K2')
        # 保存
        self.wb.save('./static/test.xlsx')

    def read_xls(self):
        """ 读取excel数据 """
        wb = load_workbook('./static/template.xlsx')
        ws = wb.active
        for (i, row) in enumerate(ws.rows):
            if i < 2:
                continue
            year = ws['A{0}'.format(i + 1)].value
            max = ws['B{0}'.format(i + 1)].value
            avg = ws['C{0}'.format(i + 1)].value
            # print("%d, %d, %d" % (year, max, avg))
            connn = self.get_conn()
            cursor = connn.cursor()
            # sql = "INSERT INTO `user_grade`.`score`(`year`, `max`, `avg`) VALUES ('2018', '706', '633')"
            sql = "INSERT INTO `user_grade`.`score`(`year`, `max`, `avg`) VALUES ({year}, {max}, {avg})"\
                .format(year=year, max=max, avg=avg)
            cursor.execute(sql)
            connn.autocommit(True)
            # print(connn)

    def get_conn(self):
        """ 获取mysql的连接 """
        try:
            conn = MySQLdb.connect(db='user_grade',
                                   host='localhost',
                                   user='root',
                                   passwd='ljz123',
                                   charset='utf8')
        except:
            pass
        return conn

    def export_xls(self):
        """ 将mysql数据库的数据导出至excel """
        # 获取数据库的连接
        conn = self.get_conn()
        cursor = conn.cursor()
        # 准备查询语句（如果数据量大，需要借助于分页查询limit 0,100然后limit 101,200这种）
        sql = "SELECT year, max, avg FROM user_grade.score;"
        # 查询数据
        cursor.execute(sql)
        rows = cursor.fetchall()
        # print(rows)
        # 循环写 入excel
        wb = Workbook()
        ws = wb.active
        for (i, row) in enumerate(rows):
            # ws['A{0}'.format(i + 1)] = row[0]
            # ws['B{0}'.format(i + 1)] = row[1]
            # ws['C{0}'.format(i + 1)] = row[2]
            (ws['A{0}'.format(i + 1)],
            ws['B{0}'.format(i + 1)],
            ws['C{0}'.format(i + 1)]) = row
        # 保存excel
        wb.save("./static/export.xlsx")
        # 也可以打开已有的excel文件，不从第一行开始写就是了


if __name__ == '__main__':
    client = ExcelUtils()
    # client.do_sth()
    # client.read_xls()
    client.export_xls()