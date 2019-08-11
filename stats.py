import MySQLdb
from openpyxl import load_workbook
from openpyxl.chart import AreaChart, Reference, Series


class GaokaoExport(object):
    def __init__(self):
        self.wb = load_workbook('./static/templ.xlsx')
        self.ws = self.wb.active
        self.ws.title = "北大高考统计"
        self.ws.sheet_properties.tabColor = "ff0000"

    def get_conn(self):
        try:
            conn = MySQLdb.connect(
                db = 'user_grade',
                host = 'localhost',
                user = 'root',
                password = 'ljz123',
                charset = 'utf8'
            )
        except:
            pass
        return conn

    def export_data(self):
        # 获取数据库的连接
        conn = self.get_conn()
        cursor = conn.cursor()
        # 准备查询语句
        sql = "SELECT year, max, avg FROM user_grade.score;"
        # 查询数据
        cursor.execute(sql)
        rows = cursor.fetchall()
        # 循环写 入excel
        row_id = 10
        for (i, row) in enumerate(rows):
            (self.ws['C{0}'.format(row_id)],
             self.ws['D{0}'.format(row_id)],
             self.ws['E{0}'.format(row_id)]) = row
            row_id += 1

        # 显示图表
        chart = AreaChart()
        chart.title = "统计表"
        chart.style = 13
        chart.x_axis.title = '年份'
        chart.y_axis.title = '分数'
        # 横坐标
        cats = Reference(self.ws, min_col=3, min_row=10, max_row=row_id)
        # 数据区域
        # 数据区域的min_row比横坐标的小1，是为了把第一行当作折线的名称
        data = Reference(self.ws, min_col=4, min_row=9, max_col=5, max_row=row_id)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        self.ws.add_chart(chart, "A{0}".format(row_id + 2))
        # 保存excel
        self.wb.save('./static/templ.xlsx')

if __name__ == '__main__':
    client = GaokaoExport()
    client.export_data()